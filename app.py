"""
Nifty 50 Market Analyzer
========================
- Real data from Yahoo Finance (no API key)
- Top Gainers & Losers for chosen date range
- ML ensemble prediction (RandomForest + GradientBoosting + Ridge)
- News sentiment from free RSS feeds (no API key)
- Excel report download: Gainers / Losers / Predictions / Summary
"""

import re
import io
from datetime import date, timedelta

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline

# ══════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════

STOCKS = {
    "RELIANCE":"Energy",    "TCS":"IT",              "HDFCBANK":"Banking",
    "BHARTIARTL":"Telecom", "ICICIBANK":"Banking",   "INFOSYS":"IT",
    "SBIN":"Banking",       "HINDUNILVR":"FMCG",     "ITC":"FMCG",
    "LICI":"Insurance",     "LT":"Infra",            "BAJFINANCE":"NBFC",
    "HCLTECH":"IT",         "KOTAKBANK":"Banking",   "MARUTI":"Auto",
    "AXISBANK":"Banking",   "TITAN":"Consumer",      "SUNPHARMA":"Pharma",
    "ONGC":"Energy",        "NTPC":"Power",          "ADANIENT":"Conglomerate",
    "WIPRO":"IT",           "ULTRACEMCO":"Cement",   "POWERGRID":"Power",
    "NESTLEIND":"FMCG",     "BAJAJFINSV":"NBFC",     "JSWSTEEL":"Metals",
    "TATAMOTORS":"Auto",    "TECHM":"IT",            "INDUSINDBK":"Banking",
    "TATACONSUM":"FMCG",    "COALINDIA":"Mining",    "ASIANPAINT":"Paint",
    "HINDALCO":"Metals",    "CIPLA":"Pharma",        "DRREDDY":"Pharma",
    "BPCL":"Energy",        "GRASIM":"Cement",       "ADANIPORTS":"Infra",
    "EICHERMOT":"Auto",     "HEROMOTOCO":"Auto",     "BAJAJ-AUTO":"Auto",
    "BRITANNIA":"FMCG",     "SBILIFE":"Insurance",   "APOLLOHOSP":"Healthcare",
    "DIVISLAB":"Pharma",    "HDFCLIFE":"Insurance",  "M&M":"Auto",
    "SHRIRAMFIN":"NBFC",    "BEL":"Defence",
}

SECTOR_SCORE = {
    "IT":5,"Pharma":5,"FMCG":5,"Healthcare":5,
    "Banking":4,"NBFC":4,"Insurance":4,
    "Auto":3,"Consumer":3,"Cement":3,"Paint":3,"Defence":3,
    "Energy":2,"Power":2,"Infra":2,"Telecom":2,"Conglomerate":2,
    "Metals":1,"Mining":1,
}

FREE_RSS = [
    "https://economictimes.indiatimes.com/markets/stocks/rssfeeds/2146842.cms",
    "https://www.moneycontrol.com/rss/marketreports.xml",
    "https://feeds.content.dowjones.io/public/rss/mw_topstories",
]

POS = {"surge","rally","gain","rise","jump","profit","growth","strong","beat",
       "record","bull","upgrade","outperform","recovery","buy","boost","exceed",
       "robust","momentum","green","high","up"}
NEG = {"fall","drop","crash","loss","decline","weak","miss","cut","downgrade",
       "bear","sell","low","down","risk","concern","uncertainty","negative",
       "slump","plunge","debt","inflation","recession","fear","red","trouble"}

# ══════════════════════════════════════════════════════════════════════
#  1.  DATA FETCH
# ══════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=3600, show_spinner=False)
def _fetch(symbol: str, start: str, end: str) -> pd.DataFrame:
    """Single cached yfinance call. end = to_date+1 (Yahoo end is exclusive)."""
    try:
        df = yf.download(f"{symbol}.NS", start=start, end=end,
                         auto_adjust=True, progress=False)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        # Hard-clip to avoid any overflow rows
        df = df[df.index <= pd.Timestamp(end) - pd.Timedelta(days=1)]
        return df
    except Exception:
        return pd.DataFrame()


def _stats(symbol: str, sector: str, df: pd.DataFrame) -> dict | None:
    """Compute every feature we need from a clean OHLCV DataFrame."""
    if df.empty or len(df) < 3:
        return None

    cl  = df["Close"].astype(float)
    hi  = df["High"].astype(float)
    lo  = df["Low"].astype(float)
    vol = df["Volume"].astype(float)

    first_c = float(cl.iloc[0])
    last_c  = float(cl.iloc[-1])
    if first_c == 0:
        return None

    p_high = float(hi.max())
    p_low  = float(lo.min())
    avg_v  = float(vol.mean())
    last_v = float(vol.iloc[-1])

    # ── Core return ───────────────────────────────────────────────────
    chg = round((last_c - first_c) / first_c * 100, 2)

    # ── RSI 14 ────────────────────────────────────────────────────────
    rsi = 50.0
    if len(cl) >= 15:
        d     = cl.diff().dropna()
        gain  = d.clip(lower=0).rolling(14).mean().iloc[-1]
        loss  = (-d).clip(lower=0).rolling(14).mean().iloc[-1]
        rsi   = round(100 - 100/(1 + gain/loss), 1) if loss > 0 else 100.0

    # ── MACD 12/26/9 ──────────────────────────────────────────────────
    ema12   = cl.ewm(span=12, adjust=False).mean()
    ema26   = cl.ewm(span=26, adjust=False).mean()
    macd_v  = float((ema12 - ema26).iloc[-1])
    sig_v   = float((ema12 - ema26).ewm(span=9, adjust=False).mean().iloc[-1])
    macd_x  = 1 if macd_v > sig_v else -1

    # ── Bollinger band position ────────────────────────────────────────
    bb_pos = 50.0
    if len(cl) >= 20:
        mid   = cl.rolling(20).mean().iloc[-1]
        std   = cl.rolling(20).std().iloc[-1]
        if std > 0:
            bb_pos = round((last_c - (mid - 2*std)) / (4*std) * 100, 1)

    # ── Position in period range ──────────────────────────────────────
    rng    = p_high - p_low
    pos_rng = round((last_c - p_low) / rng * 100, 1) if rng > 0 else 50.0

    # ── Volatility & momentum ─────────────────────────────────────────
    rets = cl.pct_change().dropna()
    volat = round(float(rets.std() * 100), 2)
    mom5  = round(float((cl.iloc[-1] - cl.iloc[-min(5,len(cl))]) /
                         cl.iloc[-min(5,len(cl))] * 100), 2)

    vol_ratio = round(last_v / avg_v, 2) if avg_v > 0 else 1.0

    return dict(
        symbol=symbol, sector=sector,
        period_high=round(p_high,2), period_low=round(p_low,2),
        first_close=round(first_c,2), last_close=round(last_c,2),
        change_pct=chg, rsi=rsi,
        macd=round(macd_v,2), macd_cross=macd_x, bb_pos=bb_pos,
        pos_in_range=pos_rng, volatility=volat,
        mom5=mom5, vol_ratio=vol_ratio,
        avg_volume=int(avg_v), last_volume=int(last_v),
    )


def fetch_all(from_d: date, to_d: date, prog) -> list[dict]:
    end = str(to_d + timedelta(days=1))
    start = str(from_d)
    out = []
    for i, (sym, sec) in enumerate(STOCKS.items()):
        prog.progress((i+1)/len(STOCKS), text=f"Fetching {sym}.NS …")
        row = _stats(sym, sec, _fetch(sym, start, end))
        if row:
            out.append(row)
    return out


# ══════════════════════════════════════════════════════════════════════
#  2.  NEWS SENTIMENT  (free RSS, no API key)
# ══════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=1800, show_spinner=False)
def _sentiment(symbols: tuple) -> dict:
    import requests
    from bs4 import BeautifulSoup
    headlines = []
    for url in FREE_RSS:
        try:
            r = requests.get(url, timeout=5, headers={"User-Agent":"Mozilla/5.0"})
            soup = BeautifulSoup(r.content, "lxml-xml")
            for tag in soup.find_all(["title","description"])[:60]:
                if tag.text:
                    headlines.append(tag.text.lower())
        except Exception:
            pass

    result = {}
    for sym in symbols:
        rel = [h for h in headlines if sym.lower() in h]
        if not rel:
            result[sym] = 0.0
            continue
        words = re.findall(r'\b\w+\b', " ".join(rel))
        p = sum(1 for w in words if w in POS)
        n = sum(1 for w in words if w in NEG)
        result[sym] = round((p-n)/(p+n), 2) if (p+n) > 0 else 0.0
    return result


# ══════════════════════════════════════════════════════════════════════
#  3.  ML ENSEMBLE  (RandomForest + GradientBoosting + Ridge)
# ══════════════════════════════════════════════════════════════════════

def _build_features(stats: list[dict]) -> np.ndarray:
    rows = []
    for s in stats:
        rows.append([
            s["rsi"],
            s["macd_cross"],
            s["bb_pos"],
            s["pos_in_range"],
            s["change_pct"],
            s["mom5"],
            s["vol_ratio"],
            s["volatility"],
            SECTOR_SCORE.get(s["sector"], 2),
        ])
    return np.array(rows, dtype=float)


def _build_target(stats: list[dict]) -> np.ndarray:
    """
    Craft a target score from domain knowledge:
    lower RSI + lower range position + more drop + higher vol ratio = more attractive.
    Models learn to generalise this across the feature space.
    """
    y = []
    for s in stats:
        t = (
            (70 - s["rsi"])           * 0.30 +
            (50 - s["pos_in_range"])  * 0.25 +
            (-s["change_pct"])        * 0.15 +
            s["vol_ratio"]            * 0.10 +
            (20 - s["volatility"])    * 0.10 +
            s["macd_cross"]           * 5.0  +
            SECTOR_SCORE.get(s["sector"],2) * 2.0
        )
        y.append(t)
    return np.array(y, dtype=float)


def ml_predict(stats: list[dict]) -> list[dict]:
    if len(stats) < 5:
        for s in stats:
            s.update(ml_score=50.0, sentiment=0.0, final_score=50.0,
                     signal="🟠 HOLD", sig_color="#f59e0b")
        return stats

    X = _build_features(stats)
    y = _build_target(stats)

    # Three models
    rf = Pipeline([("sc", StandardScaler()),
                   ("m", RandomForestRegressor(n_estimators=300, max_depth=6,
                                               random_state=42, n_jobs=1))])
    gb = Pipeline([("sc", StandardScaler()),
                   ("m", GradientBoostingRegressor(n_estimators=300, max_depth=4,
                                                   learning_rate=0.04,
                                                   subsample=0.8, random_state=42))])
    rg = Pipeline([("sc", StandardScaler()), ("m", Ridge(alpha=1.0))])

    rf.fit(X, y); gb.fit(X, y); rg.fit(X, y)

    # Weighted ensemble: RF 40 + GB 40 + Ridge 20
    raw = 0.40*rf.predict(X) + 0.40*gb.predict(X) + 0.20*rg.predict(X)

    # Normalise to 0–100
    mn, mx = raw.min(), raw.max()
    ml_scores = (raw - mn)/(mx - mn)*100 if mx > mn else np.full(len(stats), 50.0)

    # Sentiment
    sent = _sentiment(tuple(s["symbol"] for s in stats))

    enriched = []
    for s, ml_sc in zip(stats, ml_scores):
        se  = sent.get(s["symbol"], 0.0)
        # final = 80% ML + 20% sentiment boost
        final = float(np.clip(ml_sc*0.80 + se*10 + 10, 0, 100))
        final = round(final, 1)

        if   final >= 72: sig, col = "🟢 STRONG BUY", "#10b981"
        elif final >= 55: sig, col = "🟡 BUY",         "#34d399"
        elif final >= 35: sig, col = "🟠 HOLD",        "#f59e0b"
        else:             sig, col = "🔴 AVOID",       "#ef4444"

        enriched.append({**s,
                         "ml_score": round(float(ml_sc),1),
                         "sentiment": se,
                         "final_score": final,
                         "signal": sig,
                         "sig_color": col})
    return enriched


# ══════════════════════════════════════════════════════════════════════
#  4.  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="f0f0f0", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    s = Side(style="thin", color="1a1a1a")
    return Border(left=s, right=s, top=s, bottom=s)
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left",   vertical="center")

def _hdr_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill("00e5a0"); cell.font = _font(True,"0D0D0D",11)
        cell.border = _border(); cell.alignment = _ctr()

def _data_row(ws, row, ncols, bg="1a1a2e"):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg); cell.font = _font()
        cell.border = _border(); cell.alignment = _ctr()

def _title(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill("0D0D0D"); c.font = Font(bold=True, color="00e5a0", size=14, name="Calibri")
    c.alignment = _ctr(); ws.row_dimensions[row].height = 32

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _movers_sheet(ws, rows, title, green):
    hdrs = ["#","Symbol","Sector","Period High ₹","Period Low ₹",
            "First Close ₹","Last Close ₹","Change %","RSI(14)","Vol Ratio","Volatility %"]
    n = len(hdrs)
    _title(ws, title, n)
    ws.row_dimensions[2].height = 4
    for c,h in enumerate(hdrs,1): ws.cell(row=3,column=c,value=h)
    _hdr_row(ws, 3, n)
    for i,s in enumerate(rows,1):
        r = i+3
        chg = s["change_pct"]
        vals = [i, s["symbol"], s["sector"],
                s["period_high"], s["period_low"],
                s["first_close"], s["last_close"],
                f"{chg:+.2f}%", s["rsi"], s["vol_ratio"],
                f"{s['volatility']:.2f}%"]
        for c,v in enumerate(vals,1): ws.cell(row=r,column=c,value=v)
        _data_row(ws, r, n, "052e1a" if green else "2e0505")
        col = "10b981" if chg >= 0 else "ff5472"
        ws.cell(row=r,column=8).font = _font(True, col)
    _widths(ws, [4,14,13,14,14,14,14,11,10,12,12])
    ws.sheet_view.showGridLines = False


def _pred_sheet(ws, rows):
    hdrs = ["#","Symbol","Sector","ML Score","Sentiment",
            "Final Score","Signal","Period High ₹","Period Low ₹",
            "Last Close ₹","Change %","RSI","MACD","BB Pos %"]
    n = len(hdrs)
    _title(ws, "🤖 AI Predictions — Buy Signals", n)
    ws.row_dimensions[2].height = 4
    for c,h in enumerate(hdrs,1): ws.cell(row=3,column=c,value=h)
    _hdr_row(ws, 3, n)
    for i,s in enumerate(rows,1):
        r = i+3
        sig = s.get("signal","🟠 HOLD")
        sc  = s.get("final_score",50)
        if "STRONG BUY" in sig: bg,fc = "052e1a","10b981"
        elif "BUY"       in sig: bg,fc = "0a2a0a","34d399"
        elif "HOLD"      in sig: bg,fc = "1a1500","f59e0b"
        else:                    bg,fc = "1a0505","ef4444"
        se  = s.get("sentiment",0.0)
        vals= [i, s["symbol"], s["sector"],
               s.get("ml_score",50), f"{se:+.2f}", sc,
               sig.replace("🟢","").replace("🟡","").replace("🟠","").replace("🔴","").strip(),
               s["period_high"], s["period_low"], s["last_close"],
               f"{s['change_pct']:+.2f}%", s["rsi"],
               "Bullish" if s.get("macd_cross",0)>0 else "Bearish",
               f"{s.get('bb_pos',50):.1f}%"]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r,column=c,value=v)
            cell.fill=_fill(bg); cell.font=_font(); cell.border=_border(); cell.alignment=_ctr()
        ws.cell(row=r,column=6).font = _font(True, fc)
        ws.cell(row=r,column=7).font = _font(True, fc)
    _widths(ws, [4,14,13,11,11,12,15,14,14,13,11,10,11,12])
    ws.sheet_view.showGridLines = False


def _summary_sheet(ws, all_stats, from_d, to_d):
    _title(ws, "📊 Summary Dashboard", 4)
    ws.row_dimensions[2].height = 10
    changes = [s["change_pct"] for s in all_stats]
    tg = max(all_stats, key=lambda x: x["change_pct"])
    tl = min(all_stats, key=lambda x: x["change_pct"])
    rows = [
        ("Period",         f"{from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}","f0f0f0"),
        ("Stocks analysed",len(all_stats),"f0f0f0"),
        ("Avg market return",f"{sum(changes)/len(changes):+.2f}%","f0f0f0"),
        ("Gainers",        sum(1 for c in changes if c>0),"10b981"),
        ("Losers",         sum(1 for c in changes if c<0),"ff5472"),
        ("Top Gainer",     f"{tg['symbol']}  {tg['change_pct']:+.2f}%","10b981"),
        ("Top Loser",      f"{tl['symbol']}  {tl['change_pct']:+.2f}%","ff5472"),
        ("STRONG BUY",     sum(1 for s in all_stats if "STRONG BUY" in s.get("signal","")),"10b981"),
        ("BUY",            sum(1 for s in all_stats if s.get("signal","").strip()=="🟡 BUY"),"34d399"),
        ("HOLD",           sum(1 for s in all_stats if "HOLD" in s.get("signal","")),"f59e0b"),
        ("AVOID",          sum(1 for s in all_stats if "AVOID" in s.get("signal","")),"ef4444"),
    ]
    for i,(k,v,col) in enumerate(rows, 3):
        for c,val in enumerate([k,v],1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill=_fill("1a1a2e"); cell.border=_border()
            cell.alignment=_lft()
            cell.font = _font(c==1, "888888" if c==1 else col)
        ws.row_dimensions[i].height = 22
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32
    r = len(rows)+5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=1,
                   value="⚠ For educational purposes only. Not financial advice. Consult a SEBI-registered advisor.")
    cell.font=Font(italic=True,color="f59e0b",size=9,name="Calibri")
    cell.fill=_fill("1a1a2e"); cell.alignment=_lft()
    ws.sheet_view.showGridLines = False


def make_excel(all_stats, gainers, losers, predictions, from_d, to_d) -> bytes:
    wb = openpyxl.Workbook()
    label = f"{from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}"

    ws1 = wb.active; ws1.title = "Top Gainers"
    ws1.sheet_properties.tabColor = "10b981"
    _movers_sheet(ws1, gainers, f"📈 Top Gainers  ·  {label}", True)

    ws2 = wb.create_sheet("Top Losers")
    ws2.sheet_properties.tabColor = "ef4444"
    _movers_sheet(ws2, losers, f"📉 Top Losers  ·  {label}", False)

    ws3 = wb.create_sheet("AI Predictions")
    ws3.sheet_properties.tabColor = "00e5a0"
    _pred_sheet(ws3, predictions)

    ws4 = wb.create_sheet("Summary")
    ws4.sheet_properties.tabColor = "f59e0b"
    _summary_sheet(ws4, all_stats, from_d, to_d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════
#  5.  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Nifty 50 Analyzer", page_icon="📊", layout="wide")

st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#0a0a0a;}
[data-testid="stSidebar"]{background:#0f0f0f;}
h1,h2,h3{color:#f0f0f0;}
.metric-card{background:#111;border:1px solid #1e1e1e;border-radius:12px;
             padding:18px 20px;text-align:center;}
.top-card{background:#0d1f16;border:1px solid #10b981;border-radius:12px;padding:16px;text-align:center;}
.loss-card{background:#1f0d0d;border:1px solid #ef4444;border-radius:12px;padding:16px;text-align:center;}
.pred-card{background:#111;border:1px solid #333;border-radius:12px;padding:16px;text-align:center;}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📊 Nifty 50 Analyzer")
    st.markdown("---")
    today = date.today()

    preset = st.selectbox("Quick Preset", ["Custom","1 Week","2 Weeks","1 Month","3 Months","6 Months","YTD"])
    preset_map = {
        "1 Week": timedelta(weeks=1), "2 Weeks": timedelta(weeks=2),
        "1 Month": timedelta(days=30), "3 Months": timedelta(days=90),
        "6 Months": timedelta(days=180),
    }
    if preset == "YTD":
        def_from = date(today.year,1,1)
    elif preset in preset_map:
        def_from = today - preset_map[preset]
    else:
        def_from = today - timedelta(days=30)

    from_d = st.date_input("From", value=def_from, max_value=today-timedelta(days=2))
    to_d   = st.date_input("To",   value=today,    max_value=today)

    if from_d >= to_d:
        st.error("From must be before To"); st.stop()

    run = st.button("▶ Analyse", use_container_width=True, type="primary")

    if st.button("🔄 Refresh data", use_container_width=True):
        _fetch.clear(); _sentiment.clear()
        st.session_state.clear(); st.rerun()

    st.markdown("---")
    st.caption("Real NSE data · ML ensemble · News sentiment · No API key")


# ── Run analysis ────────────────────────────────────────────────────────────────
if run:
    prog = st.progress(0, "Starting…")
    all_stats = fetch_all(from_d, to_d, prog)
    prog.empty()

    if not all_stats:
        st.error("No data returned. Try a wider date range."); st.stop()

    with st.spinner("Running ML models + fetching news sentiment…"):
        enriched = ml_predict(all_stats)

    enriched.sort(key=lambda x: x["change_pct"], reverse=True)
    st.session_state["data"]   = enriched
    st.session_state["from_d"] = from_d
    st.session_state["to_d"]   = to_d
    st.rerun()


if "data" not in st.session_state:
    st.markdown("""
    <div style='text-align:center;padding:80px 20px'>
      <div style='font-size:56px'>📊</div>
      <h2 style='color:#333'>Select a date range and click Analyse</h2>
      <p style='color:#444'>Real NSE data · ML Prediction · News Sentiment · Excel Download</p>
    </div>""", unsafe_allow_html=True)
    st.stop()


# ── Render results ──────────────────────────────────────────────────────────────
data   = st.session_state["data"]
from_d = st.session_state["from_d"]
to_d   = st.session_state["to_d"]
label  = f"{from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}"

gainers     = sorted(data, key=lambda x: x["change_pct"], reverse=True)
losers      = sorted(data, key=lambda x: x["change_pct"])
predictions = sorted(data, key=lambda x: x["final_score"], reverse=True)

st.markdown(f"### 📊 Results · {label} · {len(data)} stocks")
st.markdown("---")

# ── Download button ─────────────────────────────────────────────────────────────
xlsx = make_excel(data, gainers[:10], losers[:10], predictions, from_d, to_d)
fname = f"Nifty50_{from_d}_{to_d}.xlsx"
st.download_button("📥 Download Full Report (Excel)",
                   data=xlsx, file_name=fname,
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   use_container_width=False)

st.markdown("---")

# ── 4 Tabs ──────────────────────────────────────────────────────────────────────
t1, t2, t3, t4 = st.tabs(["📈 Top Gainers", "📉 Top Losers", "🤖 AI Predictions", "📋 All Stocks"])


# ── Tab 1: Top Gainers ──────────────────────────────────────────────────────────
with t1:
    st.markdown(f"#### 📈 Top 10 Gainers  ·  {label}")
    top10g = gainers[:10]
    cols = st.columns(5)
    for i, s in enumerate(top10g):
        with cols[i % 5]:
            st.markdown(f"""
            <div class="top-card">
              <div style="font-size:15px;font-weight:900;color:#fff">{s['symbol']}</div>
              <div style="font-size:10px;color:#555;margin-bottom:6px">{s['sector']}</div>
              <div style="font-size:24px;font-weight:900;color:#10b981">{s['change_pct']:+.2f}%</div>
              <hr style="border-color:#1a3a28;margin:8px 0">
              <div style="font-size:10px;color:#888">
                High ₹{s['period_high']:,.0f}<br>
                Low  ₹{s['period_low']:,.0f}<br>
                Last ₹{s['last_close']:,.0f}
              </div>
            </div>""", unsafe_allow_html=True)
        if i == 4: st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    df_g = pd.DataFrame(gainers[:10])[[
        "symbol","sector","period_high","period_low","first_close","last_close","change_pct","rsi","vol_ratio"
    ]].rename(columns={
        "symbol":"Symbol","sector":"Sector","period_high":"Period High","period_low":"Period Low",
        "first_close":"First Close","last_close":"Last Close","change_pct":"Change %",
        "rsi":"RSI(14)","vol_ratio":"Vol Ratio"
    })
    st.dataframe(df_g.style.format({
        "Period High":"₹{:,.0f}","Period Low":"₹{:,.0f}",
        "First Close":"₹{:,.0f}","Last Close":"₹{:,.0f}",
        "Change %":"{:+.2f}%","RSI(14)":"{:.1f}","Vol Ratio":"{:.2f}x"
    }), use_container_width=True)


# ── Tab 2: Top Losers ───────────────────────────────────────────────────────────
with t2:
    st.markdown(f"#### 📉 Top 10 Losers  ·  {label}")
    top10l = losers[:10]
    cols = st.columns(5)
    for i, s in enumerate(top10l):
        with cols[i % 5]:
            st.markdown(f"""
            <div class="loss-card">
              <div style="font-size:15px;font-weight:900;color:#fff">{s['symbol']}</div>
              <div style="font-size:10px;color:#555;margin-bottom:6px">{s['sector']}</div>
              <div style="font-size:24px;font-weight:900;color:#ef4444">{s['change_pct']:+.2f}%</div>
              <hr style="border-color:#3a1a1a;margin:8px 0">
              <div style="font-size:10px;color:#888">
                High ₹{s['period_high']:,.0f}<br>
                Low  ₹{s['period_low']:,.0f}<br>
                Last ₹{s['last_close']:,.0f}
              </div>
            </div>""", unsafe_allow_html=True)
        if i == 4: st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    df_l = pd.DataFrame(losers[:10])[[
        "symbol","sector","period_high","period_low","first_close","last_close","change_pct","rsi","vol_ratio"
    ]].rename(columns={
        "symbol":"Symbol","sector":"Sector","period_high":"Period High","period_low":"Period Low",
        "first_close":"First Close","last_close":"Last Close","change_pct":"Change %",
        "rsi":"RSI(14)","vol_ratio":"Vol Ratio"
    })
    st.dataframe(df_l.style.format({
        "Period High":"₹{:,.0f}","Period Low":"₹{:,.0f}",
        "First Close":"₹{:,.0f}","Last Close":"₹{:,.0f}",
        "Change %":"{:+.2f}%","RSI(14)":"{:.1f}","Vol Ratio":"{:.2f}x"
    }), use_container_width=True)


# ── Tab 3: AI Predictions ───────────────────────────────────────────────────────
with t3:
    st.markdown(f"#### 🤖 AI Predictions  ·  {label}")
    st.caption("Ensemble of RandomForest + GradientBoosting + Ridge · News sentiment from RSS (no API)")

    buy_stocks = [s for s in predictions if "BUY" in s.get("signal","")][:10]

    if buy_stocks:
        st.markdown("##### 🏆 Top Buy Signals")
        cols = st.columns(5)
        for i, s in enumerate(buy_stocks[:5]):
            with cols[i]:
                sc = s["final_score"]
                sent = s["sentiment"]
                sent_col = "#10b981" if sent >= 0 else "#ef4444"
                st.markdown(f"""
                <div class="pred-card" style="border-color:{s['sig_color']}">
                  <div style="font-size:15px;font-weight:900;color:#fff">{s['symbol']}</div>
                  <div style="font-size:10px;color:#555;margin-bottom:6px">{s['sector']}</div>
                  <div style="font-size:26px;font-weight:900;color:{s['sig_color']}">{sc:.0f}<span style="font-size:12px">/100</span></div>
                  <div style="font-size:12px;color:{s['sig_color']};margin:4px 0">{s['signal']}</div>
                  <hr style="border-color:#222;margin:8px 0">
                  <div style="font-size:10px;color:#888">
                    ML {s['ml_score']:.0f} · Sent <span style="color:{sent_col}">{sent:+.2f}</span><br>
                    RSI {s['rsi']:.1f} · {s['change_pct']:+.2f}%
                  </div>
                </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Full predictions table
    df_p = pd.DataFrame(predictions)[[
        "symbol","sector","final_score","signal","ml_score","sentiment",
        "change_pct","rsi","macd_cross","bb_pos","period_high","period_low","last_close"
    ]].rename(columns={
        "symbol":"Symbol","sector":"Sector","final_score":"Score",
        "signal":"Signal","ml_score":"ML Score","sentiment":"Sentiment",
        "change_pct":"Change %","rsi":"RSI","macd_cross":"MACD",
        "bb_pos":"BB Pos %","period_high":"High","period_low":"Low","last_close":"Last"
    })

    def _color_sig(v):
        if "STRONG BUY" in str(v): return "background:#052e1a;color:#10b981"
        if "BUY"         in str(v): return "background:#0a2a0a;color:#34d399"
        if "HOLD"        in str(v): return "background:#1a1500;color:#f59e0b"
        return "background:#1a0505;color:#ef4444"

    styled = (df_p.style
        .applymap(_color_sig, subset=["Signal"])
        .format({"Score":"{:.1f}","ML Score":"{:.1f}","Sentiment":"{:+.2f}",
                 "Change %":"{:+.2f}%","RSI":"{:.1f}","BB Pos %":"{:.1f}%",
                 "High":"₹{:,.0f}","Low":"₹{:,.0f}","Last":"₹{:,.0f}"})
    )
    st.dataframe(styled, use_container_width=True, height=500)


# ── Tab 4: All Stocks ───────────────────────────────────────────────────────────
with t4:
    st.markdown(f"#### 📋 All {len(data)} Stocks  ·  {label}")
    df_all = pd.DataFrame(sorted(data, key=lambda x: x["change_pct"], reverse=True))[[
        "symbol","sector","period_high","period_low","first_close",
        "last_close","change_pct","rsi","vol_ratio","volatility","final_score","signal"
    ]].rename(columns={
        "symbol":"Symbol","sector":"Sector","period_high":"High","period_low":"Low",
        "first_close":"First","last_close":"Last","change_pct":"Change %",
        "rsi":"RSI","vol_ratio":"Vol Ratio","volatility":"Volatility %",
        "final_score":"AI Score","signal":"Signal"
    })

    def _chg_color(v):
        try:
            return "color:#10b981" if float(str(v).replace("%","").replace("+","")) >= 0 else "color:#ef4444"
        except: return ""

    styled_all = (df_all.style
        .applymap(_color_sig, subset=["Signal"])
        .applymap(_chg_color, subset=["Change %"])
        .format({"High":"₹{:,.0f}","Low":"₹{:,.0f}","First":"₹{:,.0f}","Last":"₹{:,.0f}",
                 "Change %":"{:+.2f}%","RSI":"{:.1f}","Vol Ratio":"{:.2f}x",
                 "Volatility %":"{:.2f}%","AI Score":"{:.1f}"})
    )
    st.dataframe(styled_all, use_container_width=True, height=700)
