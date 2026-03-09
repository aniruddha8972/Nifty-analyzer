"""
pipeline/report.py  —  v2: Charts + Global News + 5 sheets
────────────────────────────────────────────────────────────
Sheets:
  1. Top Gainers      — with News Headlines
  2. Top Losers       — with News Headlines
  3. AI Predictions   — with News Headlines
  4. Index Charts     — embedded Plotly PNG charts (Nifty 50/Next50/Midcap/Smallcap)
  5. Market News      — global macro headlines + stock news feed
  6. Summary Dashboard

Charts are rendered to PNG via Plotly's kaleido backend (if available)
and embedded as openpyxl images. If kaleido is not installed the sheet
still generates — cells contain placeholder text instead.
"""

import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold: bool = False, color: str = "f0f0f0", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border() -> Border:
    s = Side(style="thin", color="1a1a1a")
    return Border(left=s, right=s, top=s, bottom=s)

def _ctr() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _lft(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _fill("00e5a0")
        cell.font      = _font(True, "0D0D0D", 11)
        cell.border    = _border()
        cell.alignment = _ctr()

def _data_row(ws, row: int, ncols: int, bg: str = "1a1a2e") -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _fill(bg)
        cell.font      = _font()
        cell.border    = _border()
        cell.alignment = _ctr()

def _title(ws, text: str, ncols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = _fill("0D0D0D")
    cell.font      = Font(bold=True, color="00e5a0", size=14, name="Calibri")
    cell.alignment = _ctr()
    ws.row_dimensions[row].height = 32

def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _format_news(headlines: list) -> str:
    if not headlines:
        return "—  No relevant news found"
    lines = []
    for i, h in enumerate(headlines[:5], 1):
        h = h.strip()
        if h:
            lines.append(f"{i}. {h}")
    return "\n".join(lines) if lines else "—  No relevant news found"


# ── Sheet 1 & 2: Gainers / Losers ─────────────────────────────────────────────

def _write_movers(ws, rows: list[dict], title: str, is_gain: bool) -> None:
    hdrs = [
        "#", "Symbol", "Sector",
        "Period High ₹", "Period Low ₹", "First Close ₹", "Last Close ₹",
        "Change %", "RSI(14)", "Vol Ratio", "Volatility %",
        "Sentiment Score", "News Headlines (used in sentiment analysis)",
    ]
    n = len(hdrs)
    _title(ws, title, n)
    ws.row_dimensions[2].height = 4
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=3, column=c, value=h)
    _header_row(ws, 3, n)

    for i, s in enumerate(rows, 1):
        r   = i + 3
        chg = s["change_pct"]
        sent  = s.get("sentiment", 0.0)
        news  = _format_news(s.get("news_headlines", []))
        vals = [
            i, s["symbol"], s["sector"],
            s["period_high"], s["period_low"], s["first_close"], s["last_close"],
            f"{chg:+.2f}%", s["rsi"], s["vol_ratio"],
            f"{s['volatility']:.2f}%", f"{sent:+.2f}", news,
        ]
        bg = "052e1a" if is_gain else "2e0505"
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill   = _fill(bg)
            cell.font   = _font()
            cell.border = _border()
            if c == n:
                cell.alignment = _lft(wrap=True)
                cell.font = _font(color="c8c8d8", size=9)
            elif c == n - 1:
                col = "10b981" if sent >= 0 else "ff5472"
                cell.font = _font(True, col)
                cell.alignment = _ctr()
            else:
                cell.alignment = _ctr()
        ws.cell(row=r, column=8).font = _font(True, "10b981" if chg >= 0 else "ff5472")
        n_lines = news.count("\n") + 1
        ws.row_dimensions[r].height = max(20, min(n_lines * 16, 80))

    _col_widths(ws, [4, 14, 13, 14, 14, 14, 14, 11, 10, 11, 12, 13, 60])
    ws.sheet_view.showGridLines = False


# ── Sheet 3: AI Predictions ───────────────────────────────────────────────────

def _write_predictions(ws, rows: list[dict]) -> None:
    hdrs = [
        "#", "Symbol", "Sector",
        "ML Score", "Sentiment Score", "Final Score", "Signal",
        "Pred 10d Return %",
        "Period High ₹", "Period Low ₹", "Last Close ₹",
        "Change %", "RSI", "MACD", "BB Pos %",
        "News Headlines (used in sentiment analysis)",
    ]
    n = len(hdrs)
    _title(ws, "🤖  AI Predictions — Buy Signals", n)
    ws.row_dimensions[2].height = 4
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=3, column=c, value=h)
    _header_row(ws, 3, n)

    for i, s in enumerate(rows, 1):
        r   = i + 3
        sig = s.get("signal", "🟠 HOLD")
        sc  = s.get("final_score", 50)
        se  = s.get("sentiment", 0.0)
        pr  = s.get("predicted_return", 0.0)
        news = _format_news(s.get("news_headlines", []))

        if   "STRONG BUY" in sig: bg, fc = "052e1a", "10b981"
        elif "BUY"         in sig: bg, fc = "0a2a0a", "34d399"
        elif "HOLD"        in sig: bg, fc = "1a1500", "f59e0b"
        else:                      bg, fc = "1a0505", "ef4444"

        sent_col = "10b981" if se >= 0 else "ff5472"
        pred_col = "10b981" if pr >= 0 else "ff5472"
        sig_clean = (sig.replace("🟢","").replace("🟡","")
                        .replace("🟠","").replace("🔴","").strip())

        vals = [
            i, s["symbol"], s["sector"],
            s.get("ml_score", 50), f"{se:+.2f}", sc, sig_clean,
            f"{pr:+.2f}%",
            s["period_high"], s["period_low"], s["last_close"],
            f"{s['change_pct']:+.2f}%", s["rsi"],
            "Bullish" if s.get("macd_cross", 0) > 0 else "Bearish",
            f"{s.get('bb_pos', 50):.1f}%", news,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill      = _fill(bg)
            cell.font      = _font()
            cell.border    = _border()
            cell.alignment = _ctr()

        ws.cell(row=r, column=5).font  = _font(True, sent_col)
        ws.cell(row=r, column=6).font  = _font(True, fc)
        ws.cell(row=r, column=7).font  = _font(True, fc)
        ws.cell(row=r, column=8).font  = _font(True, pred_col)
        news_cell = ws.cell(row=r, column=n)
        news_cell.alignment = _lft(wrap=True)
        news_cell.font      = _font(color="c8c8d8", size=9)
        n_lines = news.count("\n") + 1
        ws.row_dimensions[r].height = max(20, min(n_lines * 16, 80))

    _col_widths(ws, [4, 14, 13, 11, 13, 12, 15, 14, 14, 14, 13, 11, 10, 11, 12, 65])
    ws.sheet_view.showGridLines = False


# ── Sheet 4: Index Charts ─────────────────────────────────────────────────────

def _make_chart_png(ticker: str, name: str, period: str = "1y") -> bytes | None:
    """Render a Plotly candlestick chart to PNG bytes. Returns None if kaleido unavailable."""
    try:
        import yfinance as yf
        import pandas as pd
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        df = yf.download(ticker, period=period, interval="1wk",
                         auto_adjust=True, progress=False)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.columns = [str(c).strip() for c in df.columns]
        if df.empty or "Close" not in df.columns:
            return None

        cl  = df["Close"].squeeze().astype(float)
        op  = df["Open"].squeeze().astype(float)
        hi  = df["High"].squeeze().astype(float)
        lo  = df["Low"].squeeze().astype(float)
        ma20 = cl.rolling(20).mean()

        last  = float(cl.iloc[-1])
        first = float(cl.iloc[0])
        chg   = (last - first) / first * 100
        chg_col = "#00e5a0" if chg >= 0 else "#ff4560"

        fig = make_subplots(rows=1, cols=1)
        fig.add_trace(go.Candlestick(
            x=df.index, open=op, high=hi, low=lo, close=cl,
            name=name,
            increasing_line_color="#00e5a0", decreasing_line_color="#ff4560",
            increasing_fillcolor="#00e5a0", decreasing_fillcolor="#ff4560",
        ))
        if not ma20.isna().all():
            fig.add_trace(go.Scatter(
                x=df.index, y=ma20,
                line=dict(color="#4c8eff", width=1.5),
                name="MA 20",
            ))

        fig.update_layout(
            title=dict(
                text=f"{name}  {last:,.0f}  <span style='color:{chg_col}'>"
                     f"({chg:+.2f}%)</span>",
                font=dict(color="#e0e0ff", size=13),
            ),
            height=320, width=700,
            plot_bgcolor="#05050d", paper_bgcolor="#0a0a14",
            font=dict(family="Arial", color="#8888aa", size=9),
            xaxis_rangeslider_visible=False,
            showlegend=False,
            margin=dict(l=40, r=20, t=40, b=30),
        )
        fig.update_xaxes(gridcolor="#111120", linecolor="#111120")
        fig.update_yaxes(gridcolor="#111120", linecolor="#111120")

        return fig.to_image(format="png", scale=1.8)

    except Exception:
        return None


def _write_index_charts(ws, from_d: date, to_d: date) -> None:
    """Embed Nifty index charts as images in the sheet."""
    from openpyxl.drawing.image import Image as XlImage

    _title(ws, "📈  Nifty Index Charts  ·  1-Year Weekly OHLCV", 8)
    ws.row_dimensions[1].height = 32

    INDICES = [
        ("^NSEI",    "Nifty 50"),
        ("^NSMIDCP", "Nifty Next 50"),
        ("^CRSMID",  "Nifty Midcap 150"),
        ("^CRSLDX",  "Nifty Smallcap 250"),
    ]

    # 2 columns × 2 rows layout
    positions = ["B3", "J3", "B33", "J33"]
    row_h_start = [3, 33]

    kaleido_ok = False
    try:
        import plotly.io as pio
        pio.kaleido.scope.default_format = "png"
        kaleido_ok = True
    except Exception:
        pass

    for idx, (ticker, name) in enumerate(INDICES):
        anchor = positions[idx]
        if kaleido_ok:
            png = _make_chart_png(ticker, name, period="1y")
            if png:
                img = XlImage(io.BytesIO(png))
                img.width  = 480
                img.height = 220
                ws.add_image(img, anchor)
            else:
                # fallback cell
                r = 3 + (idx // 2) * 30
                c = 2 + (idx % 2) * 8
                cell = ws.cell(row=r, column=c,
                               value=f"{name} — chart unavailable (data fetch error)")
                cell.fill = _fill("1a1a2e")
                cell.font = _font(color="5a5a78", size=10)
        else:
            r = 3 + (idx // 2) * 30
            c = 2 + (idx % 2) * 8
            cell = ws.cell(row=r, column=c,
                           value=f"{name}\n(Install kaleido for chart images: pip install kaleido)")
            cell.fill      = _fill("1a1a2e")
            cell.font      = _font(color="5a5a78", size=10)
            cell.alignment = _lft(wrap=True)
            ws.row_dimensions[r].height = 40

    # Also write a text summary table below charts
    summary_row = 65
    ws.cell(row=summary_row, column=1, value="Index").font = _font(True, "00e5a0")
    ws.cell(row=summary_row, column=2, value="Last").font  = _font(True, "00e5a0")
    ws.cell(row=summary_row, column=3, value="1Y Change %").font = _font(True, "00e5a0")
    for c in range(1, 4):
        ws.cell(row=summary_row, column=c).fill   = _fill("0D0D0D")
        ws.cell(row=summary_row, column=c).border = _border()
        ws.cell(row=summary_row, column=c).alignment = _ctr()

    try:
        import yfinance as yf, pandas as pd
        for i, (ticker, name) in enumerate(INDICES):
            r = summary_row + 1 + i
            try:
                df = yf.download(ticker, period="1y", interval="1wk",
                                 auto_adjust=True, progress=False)
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = df.columns.get_level_values(0)
                cl = df["Close"].squeeze().astype(float)
                last  = float(cl.iloc[-1])
                first = float(cl.iloc[0])
                chg   = (last - first) / first * 100
                chg_col = "10b981" if chg >= 0 else "ff5472"
            except Exception:
                last, chg, chg_col = 0, 0, "5a5a78"

            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                cell.fill      = _fill("1a1a2e")
                cell.border    = _border()
                cell.alignment = _ctr()
                cell.font      = _font()
            ws.cell(row=r, column=1).value = name
            ws.cell(row=r, column=2).value = f"{last:,.2f}" if last else "N/A"
            ws.cell(row=r, column=3).value = f"{chg:+.2f}%" if last else "N/A"
            ws.cell(row=r, column=3).font  = _font(True, chg_col)
    except Exception:
        pass

    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.sheet_view.showGridLines = False


# ── Sheet 5: Market News ──────────────────────────────────────────────────────

def _write_news_sheet(ws, all_stats: list[dict], from_d: date, to_d: date) -> None:
    """Two sections: Global macro news + per-stock news feed."""
    _title(ws, "📰  Market News Feed  ·  Global Macro + Stock Headlines", 6)
    ws.row_dimensions[1].height = 32

    # ── Section A: Global sentiment ───────────────────────────────────
    try:
        from backend.sentiment import fetch_global_sentiment
        g = fetch_global_sentiment()
    except Exception:
        g = None

    r = 3
    ws.cell(row=r, column=1, value="GLOBAL MARKET SENTIMENT")
    ws.cell(row=r, column=1).font      = _font(True, "00e5a0", 12)
    ws.cell(row=r, column=1).fill      = _fill("0D0D0D")
    ws.cell(row=r, column=1).alignment = _lft()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    if g:
        mood_col = "10b981" if g["overall_score"] > 0.2 else "ff4560" if g["overall_score"] < -0.2 else "f59e0b"
        meta_rows = [
            ("Mood",          g["mood"],                            mood_col),
            ("Overall Score", f"{g['overall_score']:+.3f}",        mood_col),
            ("India Score",   f"{g['india_score']:+.3f}",          mood_col),
            ("World Score",   f"{g['world_score']:+.3f}",          mood_col),
            ("Confidence",    f"{g['confidence']:.0%}",            "e0e0ff"),
            ("Articles",      f"{g['n_articles']} in last 48h",    "e0e0ff"),
        ]
        for k, v, col in meta_rows:
            for c in range(1, 3):
                cell = ws.cell(row=r, column=c)
                cell.fill   = _fill("1a1a2e")
                cell.border = _border()
                cell.alignment = _lft()
            ws.cell(row=r, column=1).value = k
            ws.cell(row=r, column=1).font  = _font(False, "5a5a78")
            ws.cell(row=r, column=2).value = v
            ws.cell(row=r, column=2).font  = _font(True, col)
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="Global Headlines (last 48h)")
        ws.cell(row=r, column=1).font  = _font(True, "00e5a0", 11)
        ws.cell(row=r, column=1).fill  = _fill("0D0D0D")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

        # Header
        hdrs = ["#", "Headline", "Source", "Published", "Score", "India?"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = _fill("00e5a0"); cell.font = _font(True, "0D0D0D", 10)
            cell.border = _border(); cell.alignment = _ctr()
        r += 1

        for i, item in enumerate(g.get("headlines", [])[:25], 1):
            sc  = item.get("score", 0.0)
            pub = item.get("pub_dt")
            ts  = pub.strftime("%d %b %H:%M") if pub else ""
            sc_col = "10b981" if sc > 0 else "ff5472" if sc < 0 else "5a5a78"
            vals = [
                i,
                item.get("title", ""),
                item.get("source", ""),
                ts,
                f"{sc:+.1f}",
                "Yes" if item.get("is_india") else "No",
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill      = _fill("0e0e1c")
                cell.font      = _font(color="c0c0d8", size=9)
                cell.border    = _border()
                cell.alignment = _lft(wrap=True) if c == 2 else _ctr()
            ws.cell(row=r, column=5).font = _font(True, sc_col, 9)
            ws.row_dimensions[r].height = 28
            r += 1

    r += 2

    # ── Section B: Per-stock news ─────────────────────────────────────
    ws.cell(row=r, column=1, value="STOCK NEWS FEED")
    ws.cell(row=r, column=1).font  = _font(True, "00e5a0", 12)
    ws.cell(row=r, column=1).fill  = _fill("0D0D0D")
    ws.cell(row=r, column=1).alignment = _lft()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    with_news = [s for s in all_stats if s.get("news_count", 0) > 0 and s.get("news_headlines")]
    with_news.sort(key=lambda x: (x.get("news_count", 0), abs(x.get("sentiment", 0))), reverse=True)

    hdrs2 = ["Symbol", "Sector", "Sentiment", "Articles", "Signal", "Headlines"]
    for c, h in enumerate(hdrs2, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill("00e5a0"); cell.font = _font(True, "0D0D0D", 10)
        cell.border = _border(); cell.alignment = _ctr()
    r += 1

    for s in with_news:
        sent    = s.get("sentiment", 0.0)
        sc_col  = "10b981" if sent > 0.1 else "ff5472" if sent < -0.1 else "f59e0b"
        sig     = s.get("signal", "🟠 HOLD")
        sig_clean = sig.replace("🟢","").replace("🟡","").replace("🟠","").replace("🔴","").strip()
        news    = _format_news(s.get("news_headlines", []))

        vals = [
            s["symbol"], s.get("sector", ""),
            f"{sent:+.3f}", s.get("news_count", 0), sig_clean, news,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill      = _fill("0e0e1c")
            cell.font      = _font(color="c0c0d8", size=9)
            cell.border    = _border()
            cell.alignment = _lft(wrap=True) if c == 6 else _ctr()
        ws.cell(row=r, column=3).font = _font(True, sc_col, 9)
        n_lines = news.count("\n") + 1
        ws.row_dimensions[r].height = max(20, min(n_lines * 15, 90))
        r += 1

    _col_widths(ws, [14, 14, 13, 10, 16, 70])
    ws.sheet_view.showGridLines = False


# ── Sheet 6: Summary ──────────────────────────────────────────────────────────

def _write_summary(ws, all_stats: list[dict], from_d: date, to_d: date) -> None:
    _title(ws, "📊  Summary Dashboard", 4)
    ws.row_dimensions[2].height = 10

    changes  = [s["change_pct"] for s in all_stats]
    tg       = max(all_stats, key=lambda x: x["change_pct"])
    tl       = min(all_stats, key=lambda x: x["change_pct"])
    n_rows   = all_stats[0].get("training_rows", 0) if all_stats else 0
    n_feats  = all_stats[0].get("n_features",    0) if all_stats else 0

    kv_rows = [
        ("Period",             f"{from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}", "f0f0f0"),
        ("Stocks analysed",    len(all_stats),                                                   "f0f0f0"),
        ("Avg market return",  f"{sum(changes)/len(changes):+.2f}%",                             "f0f0f0"),
        ("Gainers",            sum(1 for c in changes if c > 0),                                 "10b981"),
        ("Losers",             sum(1 for c in changes if c < 0),                                 "ff5472"),
        ("Top Gainer",         f"{tg['symbol']}  {tg['change_pct']:+.2f}%",                     "10b981"),
        ("Top Loser",          f"{tl['symbol']}  {tl['change_pct']:+.2f}%",                     "ff5472"),
        ("STRONG BUY count",   sum(1 for s in all_stats if "STRONG BUY" in s.get("signal","")), "10b981"),
        ("BUY count",          sum(1 for s in all_stats if s.get("signal","").strip() == "🟡 BUY"), "34d399"),
        ("HOLD count",         sum(1 for s in all_stats if "HOLD" in s.get("signal","")),        "f59e0b"),
        ("AVOID count",        sum(1 for s in all_stats if "AVOID" in s.get("signal","")),       "ef4444"),
        ("",                   "",                                                                "333333"),
        ("ML Training rows",   f"{n_rows:,}" if n_rows else "N/A",                               "00e5a0"),
        ("ML Features",        f"{n_feats} (8 technical + 7 sentiment proxies + 2 mkt-relative)" if n_feats else "N/A", "00e5a0"),
        ("ML Target",          "Actual 10-day forward return",                                    "00e5a0"),
        ("Ensemble",           "RandomForest 40% + GradientBoosting 40% + Ridge 20%",            "00e5a0"),
    ]

    for i, (k, v, col) in enumerate(kv_rows, 3):
        for c, val in enumerate([k, v], 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill      = _fill("1a1a2e")
            cell.border    = _border()
            cell.alignment = _lft()
            cell.font      = _font(c == 1, "888888" if c == 1 else col)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55

    r = len(kv_rows) + 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=1,
                   value="⚠  For educational purposes only. "
                         "Not financial advice. Consult a SEBI-registered advisor.")
    cell.font      = Font(italic=True, color="f59e0b", size=9, name="Calibri")
    cell.fill      = _fill("1a1a2e")
    cell.alignment = _lft()
    ws.sheet_view.showGridLines = False


# ── Public entry point ─────────────────────────────────────────────────────────

def generate(
    all_stats:   list[dict],
    gainers:     list[dict],
    losers:      list[dict],
    predictions: list[dict],
    from_d:      date,
    to_d:        date,
) -> bytes:
    """Build and return the full Excel workbook as bytes (6 sheets)."""
    wb    = openpyxl.Workbook()
    label = f"{from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}"

    ws1 = wb.active
    ws1.title = "Top Gainers"
    ws1.sheet_properties.tabColor = "10b981"
    _write_movers(ws1, gainers, f"📈  Top Gainers  ·  {label}", True)

    ws2 = wb.create_sheet("Top Losers")
    ws2.sheet_properties.tabColor = "ef4444"
    _write_movers(ws2, losers, f"📉  Top Losers  ·  {label}", False)

    ws3 = wb.create_sheet("AI Predictions")
    ws3.sheet_properties.tabColor = "00e5a0"
    _write_predictions(ws3, predictions)

    ws4 = wb.create_sheet("Index Charts")
    ws4.sheet_properties.tabColor = "4c8eff"
    _write_index_charts(ws4, from_d, to_d)

    ws5 = wb.create_sheet("Market News")
    ws5.sheet_properties.tabColor = "f59e0b"
    _write_news_sheet(ws5, all_stats, from_d, to_d)

    ws6 = wb.create_sheet("Summary")
    ws6.sheet_properties.tabColor = "888888"
    _write_summary(ws6, all_stats, from_d, to_d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
