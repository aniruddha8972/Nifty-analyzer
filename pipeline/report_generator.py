"""
pipeline/report_generator.py
─────────────────────────────
Generates a fully-formatted multi-sheet Excel report (.xlsx)
using openpyxl with colour coding, borders, and merged cells.

Called by the Streamlit app when user clicks "Export Excel Report".
Returns raw bytes so Streamlit can serve as a download button.
"""

import io
from datetime import date, datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from backend.data_engine import (
    StockData, get_date_range_label, trading_days_estimate
)
from backend.ai_model import StockAnalysis


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "header_bg":    "1E3A5F",
    "header_fg":    "FFFFFF",
    "gain_bg":      "064E3B",
    "gain_fg":      "00E5A0",
    "loss_bg":      "4C0519",
    "loss_fg":      "FF5472",
    "alt_row":      "0F1117",
    "title_bg":     "0A0B0F",
    "title_fg":     "00E5A0",
    "subhdr_bg":    "111827",
    "subhdr_fg":    "A78BFA",
    "strong_buy":   "10B981",
    "buy":          "34D399",
    "hold":         "F59E0B",
    "sell":         "F87171",
    "strong_sell":  "EF4444",
    "risk_low":     "00E5A0",
    "risk_med":     "F59E0B",
    "risk_high":    "FF5472",
    "white":        "FFFFFF",
    "dark":         "07080D",
    "mid":          "1F2937",
    "light_text":   "D1D5DB",
    "muted":        "6B7280",
}

REC_COLORS = {
    "STRONG BUY":  C["strong_buy"],
    "BUY":         C["buy"],
    "HOLD":        C["hold"],
    "SELL":        C["sell"],
    "STRONG SELL": C["strong_sell"],
}

RISK_COLORS = {
    "Low":      C["risk_low"],
    "Medium":   C["risk_med"],
    "Med-High": C["risk_high"],
    "High":     C["risk_high"],
}


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color: str = "FFFFFF", bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)

def _border() -> Border:
    s = Side(style="thin", color="1F2937")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _apply(cell, bg=None, fg=None, bold=False, size=10, h_align="left", wrap=False):
    if bg:  cell.fill      = _fill(bg)
    if fg:  cell.font      = _font(fg, bold=bold, size=size)
    cell.alignment         = _align(h_align, wrap=wrap)
    cell.border            = _border()


# ── Sheet builders ────────────────────────────────────────────────────────────
STOCK_COLS = [
    ("#",           4),
    ("Symbol",     12),
    ("Sector",     14),
    ("Open (₹)",   11),
    ("Close (₹)",  11),
    ("Chg %",      10),
    ("High (₹)",   11),
    ("Low (₹)",    11),
    ("Volume",     13),
    ("Avg Volume", 13),
    ("MCap (₹B)",  12),
    ("P/E",         8),
    ("52W High",   11),
    ("52W Low",    11),
    ("RSI",         8),
    ("Beta",        7),
    ("Div %",       9),
    ("AI Score",   10),
    ("Rec",        14),
    ("Risk",       10),
    ("Signals",    55),
]


def _write_stock_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    title: str,
    stocks: List[StockData],
    analyses: Dict[str, StockAnalysis],
    date_label: str,
    days: int,
    sheet_type: str = "neutral",   # "gain" | "loss" | "ai"
) -> None:
    ws = wb.create_sheet(sheet_name)

    # ── Title row
    ws.merge_cells(f"A1:{get_column_letter(len(STOCK_COLS))}1")
    ws["A1"] = title
    ws["A1"].fill      = _fill(C["title_bg"])
    ws["A1"].font      = Font(name="Calibri", color=C["title_fg"], bold=True, size=14)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # ── Sub-header row
    td = trading_days_estimate(days)
    ws.merge_cells(f"A2:{get_column_letter(len(STOCK_COLS))}2")
    ws["A2"] = f"Period: {date_label}   |   {days} calendar days · ~{td} trading days   |   Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    ws["A2"].fill      = _fill(C["mid"])
    ws["A2"].font      = Font(name="Calibri", color=C["muted"], size=9)
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 16

    # ── Blank separator
    ws.row_dimensions[3].height = 4

    # ── Column headers
    for col_idx, (col_name, col_w) in enumerate(STOCK_COLS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill      = _fill(C["header_bg"])
        cell.font      = Font(name="Calibri", color=C["header_fg"], bold=True, size=10)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[4].height = 18

    # ── Data rows
    for row_i, s in enumerate(stocks):
        a   = analyses[s.symbol]
        row = 5 + row_i
        alt = row_i % 2 == 1

        row_bg = C["alt_row"] if alt else C["dark"]

        values = [
            row_i + 1,
            s.symbol,
            s.sector,
            s.open_price,
            s.close_price,
            f"{s.chg_pct:+.2f}%",
            s.high,
            s.low,
            s.volume,
            s.avg_volume,
            s.mkt_cap_b,
            s.pe_ratio,
            s.week52_high,
            s.week52_low,
            s.rsi,
            s.beta,
            s.div_yield,
            a.score,
            a.recommendation,
            a.risk_level,
            " | ".join(a.signals),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill      = _fill(row_bg)
            cell.alignment = _align("center" if col_idx in (1, 6, 12, 15, 16, 17, 18, 20) else "left",
                                    wrap=(col_idx == 21))
            cell.border    = _border()

            # Colour-coded columns
            if col_idx == 2:   # Symbol
                cell.font = Font(name="Courier New", color=C["white"], bold=True, size=10)
            elif col_idx == 6:  # Chg %
                fg = C["gain_fg"] if s.chg_pct >= 0 else C["loss_fg"]
                cell.font = Font(name="Calibri", color=fg, bold=True, size=10)
            elif col_idx == 15:  # RSI
                if   s.rsi < 35: fg = C["gain_fg"]
                elif s.rsi > 70: fg = C["loss_fg"]
                else:            fg = C["light_text"]
                cell.font = Font(name="Calibri", color=fg, bold=True, size=10)
            elif col_idx == 19:  # Rec
                rc = REC_COLORS.get(a.recommendation, C["hold"])
                cell.fill = _fill(rc + "33")
                cell.font = Font(name="Calibri", color=rc, bold=True, size=10)
            elif col_idx == 20:  # Risk
                rc = RISK_COLORS.get(a.risk_level, C["risk_med"])
                cell.font = Font(name="Calibri", color=rc, bold=True, size=10)
            else:
                cell.font = Font(name="Calibri", color=C["light_text"], size=10)

        ws.row_dimensions[row].height = 16

    # Freeze panes below headers
    ws.freeze_panes = "A5"


def _write_summary_sheet(
    wb: openpyxl.Workbook,
    gainers: List[StockData],
    losers: List[StockData],
    combined: List[StockData],
    analyses: Dict[str, StockAnalysis],
    from_date: date,
    to_date: date,
) -> None:
    ws = wb.create_sheet("Summary Dashboard")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    label = get_date_range_label(from_date, to_date)
    days  = (to_date - from_date).days
    td    = trading_days_estimate(days)

    avg_g = sum(s.chg_pct for s in gainers) / len(gainers)
    avg_l = sum(s.chg_pct for s in losers)  / len(losers)
    best  = max(combined, key=lambda s: analyses[s.symbol].score)

    by_rec = {}
    for rec in ["STRONG BUY", "BUY", "HOLD", "SELL", "STRONG SELL"]:
        names = [s.symbol for s in combined if analyses[s.symbol].recommendation == rec]
        by_rec[rec] = ", ".join(names) if names else "(none)"

    def row(label_val, data_val, label_bg=C["mid"], label_fg=C["muted"],
            data_bg=C["dark"], data_fg=C["light_text"], bold_label=False, bold_data=False):
        r = ws.max_row + 1
        c1 = ws.cell(row=r, column=1, value=label_val)
        c2 = ws.cell(row=r, column=2, value=data_val)
        c1.fill = _fill(label_bg); c1.font = Font(name="Calibri", color=label_fg, bold=bold_label, size=10)
        c2.fill = _fill(data_bg);  c2.font = Font(name="Calibri", color=data_fg,  bold=bold_data,  size=10)
        c1.alignment = _align("left"); c2.alignment = _align("left", wrap=True)
        c1.border = _border(); c2.border = _border()
        ws.row_dimensions[r].height = 16

    def section(title):
        r = ws.max_row + 1
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=f"  {title}")
        c.fill = _fill(C["subhdr_bg"])
        c.font = Font(name="Calibri", color=C["subhdr_fg"], bold=True, size=11)
        c.alignment = _align("left")
        c.border = _border()
        ws.row_dimensions[r].height = 22

    def blank():
        r = ws.max_row + 1
        for col in (1, 2):
            c = ws.cell(row=r, column=col, value="")
            c.fill = _fill(C["dark"]); c.border = _border()
        ws.row_dimensions[r].height = 8

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "  NIFTY 50 MARKET ANALYSIS — SUMMARY DASHBOARD"
    ws["A1"].fill      = _fill(C["title_bg"])
    ws["A1"].font      = Font(name="Calibri", color=C["title_fg"], bold=True, size=14)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[1].height = 30

    blank()
    section("REPORT DETAILS")
    row("Analysis Period",   label,             bold_data=True, data_fg=C["gain_fg"])
    row("From Date",         from_date.strftime("%d %b %Y"))
    row("To Date",           to_date.strftime("%d %b %Y"))
    row("Calendar Days",     days)
    row("Est. Trading Days", td)
    row("Generated On",      datetime.now().strftime("%d %b %Y  %H:%M"))
    row("Stocks Screened",   "50 (Full Nifty 50)")
    row("In This Report",    "20 (Top 10 Gainers + Top 10 Losers)")

    blank()
    section("PERFORMANCE HIGHLIGHTS")
    row("Top Gainer",        f"{gainers[0].symbol}   +{gainers[0].chg_pct:.2f}%   (₹{gainers[0].open_price} → ₹{gainers[0].close_price})", data_fg=C["gain_fg"], bold_data=True)
    row("Top Loser",         f"{losers[0].symbol}   {losers[0].chg_pct:.2f}%   (₹{losers[0].open_price} → ₹{losers[0].close_price})",   data_fg=C["loss_fg"], bold_data=True)
    row("Avg Period Gain",   f"+{avg_g:.2f}%  (Top 10 gainers)")
    row("Avg Period Loss",   f"{avg_l:.2f}%  (Top 10 losers)")
    row("AI Best Pick",      f"{best.symbol}   Score: {analyses[best.symbol].score}/100   Rec: {analyses[best.symbol].recommendation}", data_fg=C["subhdr_fg"], bold_data=True)

    blank()
    section("AI RECOMMENDATIONS")
    for rec in ["STRONG BUY", "BUY", "HOLD", "SELL", "STRONG SELL"]:
        row(rec, by_rec[rec], data_fg=REC_COLORS.get(rec, C["white"]), bold_data=True)

    blank()
    section("SCORE THRESHOLDS")
    for label_t, rng_t in [("STRONG BUY","≥ 72"),("BUY","58–71"),("HOLD","43–57"),("SELL","29–42"),("STRONG SELL","≤ 28")]:
        row(label_t, rng_t, data_fg=REC_COLORS.get(label_t, C["white"]))

    blank()
    section("SCORING FACTORS (9 FACTORS)")
    factors = [
        ("RSI",              "Oversold <35 → +15pts; Overbought >70 → -15pts"),
        ("Period Momentum",  "Scales with period length; strong gains/losses ±8–13pts"),
        ("52W Range",        "Near 52W low → +14pts; Near 52W high → -11pts"),
        ("Volume",           ">1.6× avg → +9pts; <0.55× avg → -5pts"),
        ("P/E Valuation",    "<12 → +10pts; <18 → +6pts; >35 → -4pts; >55 → -9pts"),
        ("Beta",             "<0.65 → +6pts (Low risk); >1.5 → -6pts (High risk)"),
        ("Dividend Yield",   ">2.5% → +6pts; >1% → +3pts"),
        ("Sector",           "FMCG/Pharma/IT/Healthcare defensive bonus → +3pts"),
        ("Mean Reversion",   "Period >60d AND decline >5% → +4pts"),
    ]
    for f_label, f_desc in factors:
        row(f_label, f_desc)

    blank()
    section("DISCLAIMER")
    row("⚠ WARNING",        "FOR INFORMATIONAL PURPOSES ONLY. This is NOT financial advice.", data_fg=C["loss_fg"], bold_data=True)
    row("",                  "Past performance does not guarantee future results.")
    row("",                  "Always consult a SEBI-registered investment advisor before investing.")

    ws.freeze_panes = "A2"


# ── Public API ────────────────────────────────────────────────────────────────
def generate_excel_report(
    gainers:   List[StockData],
    losers:    List[StockData],
    analyses:  Dict[str, StockAnalysis],
    from_date: date,
    to_date:   date,
) -> bytes:
    """
    Build an Excel workbook with 4 sheets and return raw bytes.
    Streamlit calls this and passes result to st.download_button().
    """
    combined  = gainers + losers
    label     = get_date_range_label(from_date, to_date)
    days      = (to_date - from_date).days

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    _write_stock_sheet(wb, "Top 10 Gainers",  "NIFTY 50 — TOP 10 GAINERS",          gainers,  analyses, label, days, "gain")
    _write_stock_sheet(wb, "Top 10 Losers",   "NIFTY 50 — TOP 10 LOSERS",           losers,   analyses, label, days, "loss")
    _write_stock_sheet(wb, "AI Analysis",     "NIFTY 50 — AI ANALYSIS (ALL 20)",    combined, analyses, label, days, "ai")
    _write_summary_sheet(wb, gainers, losers, combined, analyses, from_date, to_date)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
