"""
tests/test_backend.py
─────────────────────
Unit tests for data engine and AI scoring model.
Run: python -m pytest tests/ -v
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from datetime import date
import pytest

from backend.data_engine import (
    generate_stock, fetch_all_stocks,
    get_top_gainers, get_top_losers,
    get_date_range_label, trading_days_estimate,
    NIFTY50_SYMBOLS,
)
from backend.ai_model import analyse_stock, analyse_all, StockAnalysis


# ── Data engine tests ─────────────────────────────────────────────────────────
class TestDataEngine:

    def setup_method(self):
        self.from_date = date(2025, 1, 1)
        self.to_date   = date(2025, 1, 31)

    def test_generate_stock_returns_valid_data(self):
        s = generate_stock("RELIANCE", self.from_date, self.to_date)
        assert s.symbol == "RELIANCE"
        assert s.sector == "Energy"
        assert s.open_price  > 0
        assert s.close_price > 0
        assert s.high >= max(s.open_price, s.close_price)
        assert s.low  <= min(s.open_price, s.close_price)
        assert s.volume > 0
        assert 0  <= s.rsi  <= 100
        assert 0  <= s.beta <= 5
        assert 0  <= s.div_yield <= 10
        assert s.days == 30

    def test_generate_stock_is_deterministic(self):
        s1 = generate_stock("TCS", self.from_date, self.to_date)
        s2 = generate_stock("TCS", self.from_date, self.to_date)
        assert s1.close_price == s2.close_price
        assert s1.chg_pct     == s2.chg_pct

    def test_different_dates_give_different_results(self):
        s1 = generate_stock("TCS", date(2024,1,1), date(2024,2,1))
        s2 = generate_stock("TCS", date(2025,1,1), date(2025,2,1))
        assert s1.chg_pct != s2.chg_pct

    def test_fetch_all_stocks_returns_all_50(self):
        stocks = fetch_all_stocks(self.from_date, self.to_date)
        assert len(stocks) == 50

    def test_stocks_sorted_descending(self):
        stocks = fetch_all_stocks(self.from_date, self.to_date)
        for i in range(len(stocks) - 1):
            assert stocks[i].chg_pct >= stocks[i+1].chg_pct

    def test_top_gainers_count(self):
        stocks  = fetch_all_stocks(self.from_date, self.to_date)
        gainers = get_top_gainers(stocks)
        assert len(gainers) == 10

    def test_top_losers_count(self):
        stocks = fetch_all_stocks(self.from_date, self.to_date)
        losers = get_top_losers(stocks)
        assert len(losers) == 10

    def test_gainers_losers_no_overlap(self):
        stocks  = fetch_all_stocks(self.from_date, self.to_date)
        gainers = {s.symbol for s in get_top_gainers(stocks)}
        losers  = {s.symbol for s in get_top_losers(stocks)}
        assert gainers.isdisjoint(losers)

    def test_gainers_all_positive_or_best(self):
        stocks  = fetch_all_stocks(self.from_date, self.to_date)
        gainers = get_top_gainers(stocks)
        losers  = get_top_losers(stocks)
        min_gainer = min(g.chg_pct for g in gainers)
        max_loser  = max(l.chg_pct for l in losers)
        assert min_gainer >= max_loser

    def test_invalid_date_range_raises(self):
        with pytest.raises(ValueError):
            generate_stock("TCS", date(2025,2,1), date(2025,1,1))

    def test_date_range_label(self):
        label = get_date_range_label(date(2025,1,1), date(2025,1,31))
        assert "Jan" in label
        assert "2025" in label

    def test_trading_days_estimate(self):
        assert trading_days_estimate(7)  == 5
        assert trading_days_estimate(14) == 10
        assert trading_days_estimate(30) == 21


# ── AI model tests ─────────────────────────────────────────────────────────────
class TestAIModel:

    def setup_method(self):
        self.from_date = date(2025, 1, 1)
        self.to_date   = date(2025, 1, 31)
        all_stocks     = fetch_all_stocks(self.from_date, self.to_date)
        combined       = get_top_gainers(all_stocks) + get_top_losers(all_stocks)
        self.stocks    = combined
        self.analyses  = analyse_all(combined)

    def test_all_stocks_have_analysis(self):
        for s in self.stocks:
            assert s.symbol in self.analyses

    def test_score_in_range(self):
        for sym, a in self.analyses.items():
            assert 0 <= a.score <= 100, f"{sym} score={a.score} out of range"

    def test_rec_is_valid(self):
        valid = {"STRONG BUY", "BUY", "HOLD", "SELL", "STRONG SELL"}
        for sym, a in self.analyses.items():
            assert a.recommendation in valid, f"{sym} rec={a.recommendation!r} invalid"

    def test_risk_is_valid(self):
        valid = {"Low", "Medium", "Med-High", "High"}
        for sym, a in self.analyses.items():
            assert a.risk_level in valid, f"{sym} risk={a.risk_level!r} invalid"

    def test_rec_color_is_hex(self):
        for sym, a in self.analyses.items():
            assert a.rec_color.startswith("#"), f"{sym} color={a.rec_color!r}"
            assert len(a.rec_color) == 7

    def test_signals_is_list(self):
        for sym, a in self.analyses.items():
            assert isinstance(a.signals, list)
            assert len(a.signals) > 0, f"{sym} has no signals"

    def test_strong_buy_threshold(self):
        a = StockAnalysis.from_score("TEST", 75, "Low", ["test"])
        assert a.recommendation == "STRONG BUY"

    def test_strong_sell_threshold(self):
        a = StockAnalysis.from_score("TEST", 20, "High", ["test"])
        assert a.recommendation == "STRONG SELL"

    def test_hold_threshold(self):
        a = StockAnalysis.from_score("TEST", 50, "Medium", ["test"])
        assert a.recommendation == "HOLD"

    def test_score_clamped(self):
        a = StockAnalysis.from_score("TEST", 999, "Low", [])
        assert a.score == 100
        b = StockAnalysis.from_score("TEST", -50, "High", [])
        assert b.score == 0

    def test_analyse_all_returns_dict(self):
        result = analyse_all(self.stocks)
        assert isinstance(result, dict)
        assert len(result) == len(self.stocks)


# ── Report generator tests ────────────────────────────────────────────────────
class TestReportGenerator:

    def test_excel_report_returns_bytes(self):
        from pipeline.report_generator import generate_excel_report
        from_date = date(2025, 1, 1)
        to_date   = date(2025, 1, 31)
        all_stocks = fetch_all_stocks(from_date, to_date)
        gainers    = get_top_gainers(all_stocks)
        losers     = get_top_losers(all_stocks)
        combined   = gainers + losers
        analyses   = analyse_all(combined)

        result = generate_excel_report(gainers, losers, analyses, from_date, to_date)
        assert isinstance(result, bytes)
        assert len(result) > 5000   # should be a real xlsx file

    def test_excel_report_is_valid_xlsx(self):
        import io
        import openpyxl
        from pipeline.report_generator import generate_excel_report
        from_date = date(2025, 1, 1)
        to_date   = date(2025, 1, 31)
        all_stocks = fetch_all_stocks(from_date, to_date)
        gainers    = get_top_gainers(all_stocks)
        losers     = get_top_losers(all_stocks)
        analyses   = analyse_all(gainers + losers)

        data = generate_excel_report(gainers, losers, analyses, from_date, to_date)
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        assert "Top 10 Gainers"    in wb.sheetnames
        assert "Top 10 Losers"     in wb.sheetnames
        assert "AI Analysis"       in wb.sheetnames
        assert "Summary Dashboard" in wb.sheetnames


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
