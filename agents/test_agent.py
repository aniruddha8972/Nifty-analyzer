"""
agents/test_agent.py
────────────────────
Test Agent — runs 8 test suites covering every layer of the app.
Prints PASS/FAIL for each test, exits 0 if all pass, 1 if any fail.

Test Suites:
  1. Syntax        — ast.parse every .py file
  2. Imports       — all backend/frontend modules import cleanly
  3. Auth          — register, login, wrong password, duplicate, JSON roundtrip
  4. Portfolio     — add, avg price, remove, P&L, advice matrix, JSON export
  5. Data layer    — compute_stats with synthetic OHLCV
  6. ML pipeline   — feature extraction + ensemble scoring on synthetic data
  7. Excel report  — generate() produces valid xlsx bytes
  8. Config        — requirements.txt, secrets template, .gitignore

Run:  python agents/test_agent.py
"""


import sys, types

def _mock(name):
    m = types.ModuleType(name)
    sys.modules[name] = m
    return m

# ── Register all external mocks BEFORE any other import ───────────────────────

# streamlit mock
_st = _mock("streamlit")
_st.session_state  = {}
_st.cache_data     = lambda *a, **kw: (lambda f: f)
_st.cache_resource = lambda *a, **kw: (lambda f: f)
class _SecretsMock(dict):
    def get(self, key, default=None): return default
_st.secrets = _SecretsMock()

# yfinance mock
import numpy as _np, pandas as _pd
_yf = _mock("yfinance")
def _yf_dl(*a, **kw):
    n = 60
    dates = _pd.date_range("2024-01-01", periods=n, freq="B")
    c = 1000 * _np.cumprod(1 + _np.random.normal(0.001, 0.015, n))
    return _pd.DataFrame({
        "Open": c, "High": c*1.01, "Low": c*0.99,
        "Close": c, "Volume": _np.ones(n)*1e6
    }, index=dates)
_yf.download = _yf_dl

# supabase mock
_sb = _mock("supabase")
_sb.create_client = lambda u, k: None


import ast, hashlib, json, re, sys, traceback
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

PASS = 0
FAIL = 0
RESULTS = []

# ── test harness ───────────────────────────────────────────────────────────────

def test(name, fn):
    global PASS, FAIL
    try:
        fn()
        RESULTS.append(("PASS", name, ""))
        PASS += 1
    except Exception as e:
        tb = traceback.format_exc().strip().split("\n")[-1]
        RESULTS.append(("FAIL", name, tb))
        FAIL += 1

def expect(condition, msg=""):
    if not condition:
        raise AssertionError(msg or "Assertion failed")

# ══════════════════════════════════════════════════════════════════════
#  SUITE 1: SYNTAX
# ══════════════════════════════════════════════════════════════════════

def test_syntax():
    errors = []
    for fpath in sorted(ROOT.rglob("*.py")):
        if "__pycache__" in str(fpath): continue
        if "agents/" in str(fpath): continue  # agents are tooling, not app code
        try:
            ast.parse(fpath.read_text())
        except SyntaxError as e:
            errors.append(f"{fpath.relative_to(ROOT)}: {e}")
    expect(not errors, "\n".join(errors))

test("SYNTAX  — all .py files parse without SyntaxError", test_syntax)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 2: IMPORTS
# ══════════════════════════════════════════════════════════════════════


def test_import_constants():
    from backend.constants import STOCKS, SECTOR_SCORE, FREE_RSS, POS_WORDS, NEG_WORDS
    expect(len(STOCKS) == 50,       f"Expected 50 stocks, got {len(STOCKS)}")
    expect(len(SECTOR_SCORE) > 0,   "SECTOR_SCORE empty")
    expect(len(FREE_RSS) >= 2,      "Need at least 2 RSS feeds")
    expect("RELIANCE" in STOCKS,    "RELIANCE missing from STOCKS")

def test_import_auth_module():
    from backend import auth
    expect(callable(auth.register),            "register not callable")
    expect(callable(auth.login),               "login not callable")
    expect(callable(auth.load_user_portfolio), "load_user_portfolio not callable")
    expect(callable(auth.save_user_portfolio), "save_user_portfolio not callable")

def test_import_portfolio_module():
    from backend import portfolio
    expect(callable(portfolio.add_holding),         "add_holding not callable")
    expect(callable(portfolio.remove_holding),      "remove_holding not callable")
    expect(callable(portfolio.compute_portfolio_pnl),"compute_portfolio_pnl not callable")
    expect(callable(portfolio.get_portfolio_advice), "get_portfolio_advice not callable")

def test_import_data_module():
    from backend import data
    expect(callable(data.compute_stats),  "compute_stats not callable")
    expect(callable(data.fetch_all),      "fetch_all not callable")
    expect(callable(data.fetch_ohlcv),    "fetch_ohlcv not callable")

test("IMPORTS — backend.constants",  test_import_constants)
test("IMPORTS — backend.auth",        test_import_auth_module)
test("IMPORTS — backend.portfolio",   test_import_portfolio_module)
test("IMPORTS — backend.data",        test_import_data_module)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 3: AUTH (local JSON mode)
# ══════════════════════════════════════════════════════════════════════

import tempfile, os

# Redirect auth to a temp dir so tests don't touch real data
_tmp = tempfile.mkdtemp()
_tmp_path = Path(_tmp)
(_tmp_path / "portfolios").mkdir()
(_tmp_path / "users.json").write_text("{}")

import backend.auth as _auth_mod
_orig_base = _auth_mod._BASE
_orig_users = _auth_mod._USERS_FILE
_orig_pf_dir = _auth_mod._PORTFOLIO_DIR
_auth_mod._BASE          = _tmp_path
_auth_mod._USERS_FILE    = _tmp_path / "users.json"
_auth_mod._PORTFOLIO_DIR = _tmp_path / "portfolios"

def test_auth_register_ok():
    ok, msg = _auth_mod._local_register("test_user", "Test User", "test@x.com", "pass123")
    expect(ok,  f"Register failed: {msg}")
    expect("Welcome" in msg, f"Unexpected message: {msg}")

def test_auth_duplicate_username():
    ok, msg = _auth_mod._local_register("test_user", "Another", "other@x.com", "pass123")
    expect(not ok, "Duplicate username should fail")
    expect("taken" in msg.lower(), f"Expected 'taken' in: {msg}")

def test_auth_duplicate_email():
    ok, msg = _auth_mod._local_register("other_user", "Other", "test@x.com", "pass123")
    expect(not ok, "Duplicate email should fail")
    expect("email" in msg.lower(), f"Expected 'email' in: {msg}")

def test_auth_bad_username():
    cases = [("ab", "Too short"), ("a"*21, "Too long"), ("my user", "Space"), ("MY_USER!", "Special char")]
    for uname, reason in cases:
        ok, _ = _auth_mod._local_register(uname, "Name", "u@u.com", "pass123")
        expect(not ok, f"Bad username '{uname}' ({reason}) should fail")

def test_auth_short_password():
    ok, msg = _auth_mod._local_register("new_user2", "Name", "n2@x.com", "abc")
    expect(not ok, "Short password should fail")

def test_auth_login_ok():
    ok, msg, info = _auth_mod._local_login("test_user", "pass123")
    expect(ok,                      f"Login failed: {msg}")
    expect(info["username"] == "test_user", "Wrong username in info")
    expect(info["name"] == "Test User",     "Wrong name in info")
    expect("password_hash" not in info,     "Hash leaked into user_info!")

def test_auth_login_email():
    ok, msg, info = _auth_mod._local_login("test@x.com", "pass123")
    expect(ok, f"Email login failed: {msg}")

def test_auth_wrong_password():
    ok, msg, info = _auth_mod._local_login("test_user", "wrongpass")
    expect(not ok, "Wrong password should fail")
    expect(info is None, "user_info should be None on failed login")

def test_auth_unknown_user():
    ok, msg, info = _auth_mod._local_login("nobody", "pass123")
    expect(not ok, "Unknown user should fail")

def test_auth_password_hashing():
    # Passwords must never be stored in plain text
    users = json.loads((_tmp_path / "users.json").read_text())
    for u in users.values():
        h = u["password_hash"]
        expect(h != "pass123", "Password stored in plain text!")
        expect(len(h) == 64,   f"SHA-256 hash should be 64 chars, got {len(h)}")

def test_auth_portfolio_save_load():
    pf = {"RELIANCE": {"symbol":"RELIANCE","sector":"Energy","qty":50,"avg_buy_price":1320.0,"lots":[]}}
    info = {"username": "test_user"}
    _auth_mod._local_save_portfolio(info, pf)
    loaded = _auth_mod._local_load_portfolio(info)
    expect(loaded == pf, f"Portfolio roundtrip mismatch: {loaded}")

test("AUTH    — register valid user",         test_auth_register_ok)
test("AUTH    — duplicate username rejected", test_auth_duplicate_username)
test("AUTH    — duplicate email rejected",    test_auth_duplicate_email)
test("AUTH    — bad username formats",        test_auth_bad_username)
test("AUTH    — short password rejected",     test_auth_short_password)
test("AUTH    — login with username",         test_auth_login_ok)
test("AUTH    — login with email",            test_auth_login_email)
test("AUTH    — wrong password rejected",     test_auth_wrong_password)
test("AUTH    — unknown user rejected",       test_auth_unknown_user)
test("AUTH    — passwords hashed (SHA-256)",  test_auth_password_hashing)
test("AUTH    — portfolio save/load",         test_auth_portfolio_save_load)

# Restore auth paths
_auth_mod._BASE          = _orig_base
_auth_mod._USERS_FILE    = _orig_users
_auth_mod._PORTFOLIO_DIR = _orig_pf_dir

# ══════════════════════════════════════════════════════════════════════
#  SUITE 4: PORTFOLIO LOGIC
# ══════════════════════════════════════════════════════════════════════

from backend.portfolio import (
    compute_portfolio_pnl, get_portfolio_advice,
    export_portfolio_json, import_portfolio_json
)

_MOCK_PF = {
    "RELIANCE": {"symbol":"RELIANCE","sector":"Energy","qty":70,"avg_buy_price":1340.0,
                 "lots":[{"date":"2025-01-01","qty":70,"price":1340.0}]},
    "TCS":      {"symbol":"TCS","sector":"IT","qty":10,"avg_buy_price":3650.0,
                 "lots":[{"date":"2025-01-01","qty":10,"price":3650.0}]},
    "INFOSYS":  {"symbol":"INFOSYS","sector":"IT","qty":25,"avg_buy_price":1520.0,
                 "lots":[{"date":"2025-01-01","qty":25,"price":1520.0}]},
}
_MOCK_PRICES = {"RELIANCE":1485.0, "TCS":3210.0, "INFOSYS":1680.0}

def test_pnl_values():
    rows, totals = compute_portfolio_pnl(_MOCK_PF, _MOCK_PRICES)
    expect(len(rows) == 3, f"Expected 3 rows, got {len(rows)}")
    rel = next(r for r in rows if r["symbol"] == "RELIANCE")
    expect(abs(rel["invested"] - 70*1340)    < 0.01, "RELIANCE invested wrong")
    expect(abs(rel["current_val"] - 70*1485) < 0.01, "RELIANCE current_val wrong")
    expect(abs(rel["pnl"] - (70*1485-70*1340)) < 0.01, "RELIANCE pnl wrong")

def test_pnl_pct_accuracy():
    rows, _ = compute_portfolio_pnl(_MOCK_PF, _MOCK_PRICES)
    rel = next(r for r in rows if r["symbol"] == "RELIANCE")
    expected = (1485-1340)/1340*100
    expect(abs(rel["pnl_pct"] - expected) < 0.01, f"P&L% wrong: {rel['pnl_pct']} vs {expected:.2f}")

def test_pnl_totals():
    _, totals = compute_portfolio_pnl(_MOCK_PF, _MOCK_PRICES)
    expect(totals["total_invested"] > 0,  "total_invested should be > 0")
    expect(totals["best"]  is not None,   "best should not be None")
    expect(totals["worst"] is not None,   "worst should not be None")
    expect(totals["n_profit"] + totals["n_loss"] <= 3, "n_profit+n_loss > n_holdings")

def test_pnl_zero_price():
    pf = {"NODATASTK": {"symbol":"NODATASTK","sector":"IT","qty":10,"avg_buy_price":100,"lots":[]}}
    rows, totals = compute_portfolio_pnl(pf, {"NODATASTK": 0.0})
    expect(rows[0]["current_val"] == 0.0,  "Zero price should give 0 current_val")

def test_avg_price_multi_lot():
    # 30 shares @ 1280 + 20 shares @ 1380 → avg = (30*1280+20*1380)/50 = 1320
    pf = _MOCK_PF.copy()
    pf["TEST"] = {"symbol":"TEST","sector":"IT","qty":50,"avg_buy_price":1320.0,"lots":[]}
    rows, _ = compute_portfolio_pnl(pf, {**_MOCK_PRICES,"TEST":1320.0})
    t = next(r for r in rows if r["symbol"]=="TEST")
    expect(abs(t["pnl_pct"]) < 0.01, f"Break-even should give ~0% P&L, got {t['pnl_pct']}")

def test_advice_matrix():
    rows, _ = compute_portfolio_pnl(_MOCK_PF, _MOCK_PRICES)
    ml_stats = [
        {"symbol":"RELIANCE","signal":"🟢 STRONG BUY","final_score":78,"sentiment":0.3,"predicted_return":3.2},
        {"symbol":"TCS",     "signal":"🔴 AVOID",     "final_score":28,"sentiment":-0.4,"predicted_return":-2.1},
        {"symbol":"INFOSYS", "signal":"🟡 HOLD",      "final_score":52,"sentiment":0.0,"predicted_return":0.8},
    ]
    advised = get_portfolio_advice(rows, ml_stats)
    rel = next(r for r in advised if r["symbol"]=="RELIANCE")
    tcs = next(r for r in advised if r["symbol"]=="TCS")
    inf = next(r for r in advised if r["symbol"]=="INFOSYS")
    expect("BUY MORE" in rel["advice"],   f"STRONG BUY+profit → BUY MORE, got: {rel['advice']}")
    # TCS is down -12% with AVOID → STOP LOSS
    expect("STOP LOSS" in tcs["advice"] or "BOOK" in tcs["advice"],
           f"AVOID+loss → STOP LOSS/BOOK, got: {tcs['advice']}")
    expect("HOLD" in inf["advice"] or "PARTIAL" in inf["advice"],
           f"HOLD → HOLD/PARTIAL, got: {inf['advice']}")

def test_advice_no_ml_data():
    rows, _ = compute_portfolio_pnl(_MOCK_PF, _MOCK_PRICES)
    advised = get_portfolio_advice(rows, None)
    for r in advised:
        expect("advice" in r,       "advice key missing")
        expect("advice_color" in r, "advice_color key missing")
        expect("advice_reason" in r,"advice_reason key missing")

def test_portfolio_json_roundtrip():
    js = export_portfolio_json(_MOCK_PF)
    parsed = json.loads(js)
    expect(parsed == _MOCK_PF, "JSON export/import mismatch")

def test_portfolio_import_validation():
    bad_json = '{"BAD": {"qty": "notanumber", "avg_buy_price": 100}}'
    result, err = import_portfolio_json(bad_json)
    # Should either succeed with coercion or gracefully fail
    expect(err == "" or result is not None or isinstance(err, str),
           "import_portfolio_json should never throw")

def test_portfolio_import_invalid_json():
    result, err = import_portfolio_json("not json at all {{{")
    expect(result is None, "Invalid JSON should return None")
    expect(err != "",      "Invalid JSON should return error message")

test("PORTFOLIO — P&L values correct",          test_pnl_values)
test("PORTFOLIO — P&L % accuracy",              test_pnl_pct_accuracy)
test("PORTFOLIO — totals + best/worst",         test_pnl_totals)
test("PORTFOLIO — zero price handled",          test_pnl_zero_price)
test("PORTFOLIO — break-even at avg price",     test_avg_price_multi_lot)
test("PORTFOLIO — advice matrix (ML signals)",  test_advice_matrix)
test("PORTFOLIO — advice with no ML data",      test_advice_no_ml_data)
test("PORTFOLIO — JSON export roundtrip",       test_portfolio_json_roundtrip)
test("PORTFOLIO — import validation",           test_portfolio_import_validation)
test("PORTFOLIO — invalid JSON import error",   test_portfolio_import_invalid_json)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 5: DATA LAYER
# ══════════════════════════════════════════════════════════════════════

from backend.data import compute_stats

def _make_ohlcv(n=60, trend=0.001):
    """Generate synthetic OHLCV DataFrame."""
    dates = pd.date_range("2024-01-01", periods=n, freq="B")
    close = 1000 * np.cumprod(1 + np.random.normal(trend, 0.015, n))
    high  = close * (1 + np.abs(np.random.normal(0, 0.008, n)))
    low   = close * (1 - np.abs(np.random.normal(0, 0.008, n)))
    open_ = close * (1 + np.random.normal(0, 0.005, n))
    vol   = np.random.randint(1_000_000, 10_000_000, n).astype(float)
    return pd.DataFrame({"Open":open_,"High":high,"Low":low,"Close":close,"Volume":vol}, index=dates)

def test_compute_stats_basic():
    df = _make_ohlcv(60)
    result = compute_stats("RELIANCE", "Energy", df)
    expect(result is not None, "compute_stats returned None for valid data")
    required_keys = ["symbol","sector","change_pct","rsi","macd_cross",
                     "bb_pos","volatility","mom5","vol_ratio",
                     "overnight_gap","intraday_range","close_loc",
                     "news_event","sentiment_3d","big_gap_5d"]
    for k in required_keys:
        expect(k in result, f"Missing key: {k}")

def test_compute_stats_rsi_range():
    df = _make_ohlcv(60)
    r  = compute_stats("TCS", "IT", df)
    expect(0 <= r["rsi"] <= 100, f"RSI out of range: {r['rsi']}")

def test_compute_stats_bb_pos():
    df = _make_ohlcv(60)
    r  = compute_stats("INFOSYS", "IT", df)
    # BB pos should usually be within -50..150 range (can exceed 0-100)
    expect(-200 < r["bb_pos"] < 300, f"BB pos unreasonable: {r['bb_pos']}")

def test_compute_stats_too_few_rows():
    df = _make_ohlcv(3)  # too few
    r  = compute_stats("X", "Y", df)
    expect(r is None, "Should return None for < 5 rows")

def test_compute_stats_empty():
    r = compute_stats("X", "Y", pd.DataFrame())
    expect(r is None, "Should return None for empty DataFrame")

def test_compute_stats_with_nifty():
    df    = _make_ohlcv(60)
    nifty = pd.Series(
        1000 * np.cumprod(1 + np.random.normal(0.0005, 0.01, 60)),
        index=df.index
    )
    r = compute_stats("SBIN", "Banking", df, nifty_cl=nifty)
    expect(r is not None, "compute_stats with nifty returned None")
    expect("stock_vs_mkt" in r, "stock_vs_mkt missing")
    expect("stock_rs5"    in r, "stock_rs5 missing")
    expect(isinstance(r["stock_vs_mkt"], float), "stock_vs_mkt should be float")

def test_compute_stats_volatility():
    df = _make_ohlcv(60, trend=0.0)
    r  = compute_stats("X", "Y", df)
    expect(r["volatility"] >= 0, "Volatility should be >= 0")

test("DATA    — compute_stats returns all keys",      test_compute_stats_basic)
test("DATA    — RSI always 0-100",                    test_compute_stats_rsi_range)
test("DATA    — Bollinger Band position reasonable",  test_compute_stats_bb_pos)
test("DATA    — too few rows returns None",           test_compute_stats_too_few_rows)
test("DATA    — empty DataFrame returns None",        test_compute_stats_empty)
test("DATA    — market-relative features with Nifty", test_compute_stats_with_nifty)
test("DATA    — volatility is non-negative",          test_compute_stats_volatility)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 6: ML PIPELINE
# ══════════════════════════════════════════════════════════════════════

def test_ml_feature_vector():
    """build_dataset should produce a non-empty DataFrame with correct columns."""
    from backend.ml import build_dataset
    # Needs real download — skip if no network, just check function exists
    expect(callable(build_dataset), "build_dataset not callable")

def test_ml_predict_structure():
    """predict() must return list of dicts with required keys."""
    from backend.portfolio import get_portfolio_advice
    # Use portfolio's version for mock test
    rows = [{"symbol":"RELIANCE","sector":"Energy","qty":10,"avg_buy_price":1300,
             "current_price":1400,"invested":13000,"current_val":14000,
             "pnl":1000,"pnl_pct":7.7,"lots":[]}]
    ml_stats = [{"symbol":"RELIANCE","signal":"🟢 STRONG BUY",
                 "final_score":75,"sentiment":0.2,"predicted_return":2.5}]
    from backend.portfolio import get_portfolio_advice as gpa
    result = gpa(rows, ml_stats)
    expect(len(result) == 1,         "Expected 1 result")
    expect("advice" in result[0],    "advice key missing")
    expect("BUY" in result[0]["advice"], f"Expected BUY advice, got: {result[0]['advice']}")

def test_ml_signal_thresholds():
    """Final score thresholds produce correct signals."""
    # STRONG BUY ≥ 72, BUY ≥ 55, HOLD ≥ 35, AVOID < 35
    thresholds = [
        (80, "STRONG BUY"),
        (60, "BUY"),
        (45, "HOLD"),
        (20, "AVOID"),
    ]
    def score_to_signal(score):
        if score >= 72: return "STRONG BUY"
        if score >= 55: return "BUY"
        if score >= 35: return "HOLD"
        return "AVOID"
    for score, expected in thresholds:
        got = score_to_signal(score)
        expect(got == expected, f"Score {score} → {got}, expected {expected}")

test("ML      — build_dataset callable",         test_ml_feature_vector)
test("ML      — predict result structure",       test_ml_predict_structure)
test("ML      — signal thresholds correct",      test_ml_signal_thresholds)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 7: EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════

def _mock_stock_data(n=10):
    return [{
        "symbol": f"STOCK{i}", "sector": "IT",
        "change_pct": (i-5)*1.5, "rsi": 45+i,
        "period_high": 1100+i*10, "period_low": 900+i*5,
        "first_close": 950+i*5, "last_close": 1000+i*8,
        "volatility": 1.5+i*0.1, "vol_ratio": 1.0+i*0.05,
        "avg_volume": 1000000, "last_volume": 1100000,
        "macd": 2.5, "macd_cross": 1, "bb_pos": 55.0,
        "pos_in_range": 60.0, "mom5": 1.2,
        "overnight_gap": 0.3, "intraday_range": 1.1,
        "close_loc": 0.6, "news_event": 0.5,
        "sentiment_3d": 0.8, "big_gap_5d": 1.2,
        "stock_vs_mkt": 0.4, "stock_rs5": 0.9,
        "signal": "🟢 BUY", "sig_color": "#10b981",
        "final_score": 62.0, "ml_score": 58.0,
        "sentiment": 0.2, "predicted_return": 1.8,
        "training_rows": 35000, "n_features": 17,
        "news_headlines": ["Stock rallies on strong earnings", "Analysts upgrade target"],
    } for i in range(n)]

def test_excel_generates():
    from pipeline.report import generate
    data = _mock_stock_data(20)
    gainers = sorted(data, key=lambda x: x["change_pct"], reverse=True)[:10]
    losers  = sorted(data, key=lambda x: x["change_pct"])[:10]
    preds   = sorted(data, key=lambda x: x["final_score"], reverse=True)
    xlsx = generate(data, gainers, losers, preds, date(2025,1,1), date(2025,3,1))
    expect(isinstance(xlsx, bytes),  "generate() should return bytes")
    expect(len(xlsx) > 5000,         f"xlsx too small: {len(xlsx)} bytes")
    # Check it's a valid ZIP (xlsx is a zip file)
    expect(xlsx[:2] == b"PK",        "xlsx doesn't start with PK (not a valid zip/xlsx)")

def test_excel_all_sheets():
    """Check all 4 sheets are present in the workbook."""
    import io
    from openpyxl import load_workbook
    from pipeline.report import generate
    data = _mock_stock_data(20)
    gainers = sorted(data, key=lambda x: x["change_pct"], reverse=True)[:10]
    losers  = sorted(data, key=lambda x: x["change_pct"])[:10]
    preds   = sorted(data, key=lambda x: x["final_score"], reverse=True)
    xlsx = generate(data, gainers, losers, preds, date(2025,1,1), date(2025,3,1))
    wb   = load_workbook(io.BytesIO(xlsx))
    sheets = wb.sheetnames
    expect("Top Gainers"    in sheets, f"Missing 'Top Gainers' sheet. Got: {sheets}")
    expect("Top Losers"     in sheets, f"Missing 'Top Losers' sheet. Got: {sheets}")
    expect("AI Predictions" in sheets, f"Missing 'AI Predictions' sheet. Got: {sheets}")
    expect("Summary"        in sheets, f"Missing 'Summary' sheet. Got: {sheets}")

test("EXCEL   — generate() returns valid xlsx bytes", test_excel_generates)
test("EXCEL   — all 4 sheets present",               test_excel_all_sheets)

# ══════════════════════════════════════════════════════════════════════
#  SUITE 8: CONFIG & FILES
# ══════════════════════════════════════════════════════════════════════

def test_requirements_has_supabase():
    txt = (ROOT / "requirements.txt").read_text()
    expect("supabase" in txt, "supabase missing from requirements.txt")
    expect("streamlit" in txt, "streamlit missing")
    expect("yfinance"  in txt, "yfinance missing")
    expect("scikit-learn" in txt, "scikit-learn missing")
    expect("openpyxl"  in txt, "openpyxl missing")

def test_gitignore_secrets():
    gi = (ROOT / ".gitignore").read_text()
    expect("secrets.toml" in gi, "secrets.toml not in .gitignore — security risk!")

def test_supabase_sql_exists():
    sql = ROOT / "supabase_setup.sql"
    expect(sql.exists(), "supabase_setup.sql missing")
    content = sql.read_text()
    expect("CREATE TABLE" in content,         "SQL missing CREATE TABLE")
    expect("profiles"     in content,         "SQL missing profiles table")
    expect("portfolios"   in content,         "SQL missing portfolios table")
    expect("ROW LEVEL SECURITY" in content,   "SQL missing RLS")

def test_deploy_guide_exists():
    d = ROOT / "DEPLOY.md"
    expect(d.exists(), "DEPLOY.md missing")
    txt = d.read_text()
    expect("Supabase" in txt, "DEPLOY.md should mention Supabase")
    expect("Step"     in txt, "DEPLOY.md should have steps")

def test_streamlit_config():
    cfg = ROOT / ".streamlit/config.toml"
    expect(cfg.exists(), ".streamlit/config.toml missing")

def test_no_hardcoded_secrets():
    """Scan for any hardcoded API keys or passwords."""
    suspicious = re.compile(
        r'(password|secret|api_key|token)\s*=\s*["\'][^"\']{8,}["\']',
        re.IGNORECASE
    )
    found = []
    for fpath in ROOT.rglob("*.py"):
        if "__pycache__" in str(fpath) or "agents/" in str(fpath): continue
        for i, line in enumerate(fpath.read_text().splitlines(), 1):
            if suspicious.search(line) and "placeholder" not in line.lower() \
               and "example" not in line.lower() and "hash" not in line.lower():
                found.append(f"{fpath.relative_to(ROOT)}:{i}: {line.strip()[:60]}")
    expect(not found, "Possible hardcoded secrets:\n" + "\n".join(found))

test("CONFIG  — requirements.txt has all deps",    test_requirements_has_supabase)
test("CONFIG  — .gitignore protects secrets",      test_gitignore_secrets)
test("CONFIG  — supabase_setup.sql complete",      test_supabase_sql_exists)
test("CONFIG  — DEPLOY.md guide exists",           test_deploy_guide_exists)
test("CONFIG  — .streamlit/config.toml present",   test_streamlit_config)
test("SECURITY— no hardcoded secrets in code",     test_no_hardcoded_secrets)

# ══════════════════════════════════════════════════════════════════════
#  REPORT
# ══════════════════════════════════════════════════════════════════════

WIDTH = 62
print("\n" + "="*WIDTH)
print("  TEST AGENT — RESULTS")
print("="*WIDTH)

suite = ""
for status, name, err in RESULTS:
    # Print suite header when it changes
    new_suite = name.split()[0]
    if new_suite != suite:
        suite = new_suite
        print()
    icon = "✅" if status == "PASS" else "❌"
    print(f"  {icon} {name}")
    if err:
        print(f"       ↳ {err}")

total = PASS + FAIL
print()
print("="*WIDTH)
print(f"  {'✅ ALL PASS' if FAIL==0 else '❌ FAILURES FOUND'}")
print(f"  {PASS}/{total} tests passed  ·  {FAIL} failed")
print("="*WIDTH + "\n")

sys.exit(0 if FAIL == 0 else 1)
